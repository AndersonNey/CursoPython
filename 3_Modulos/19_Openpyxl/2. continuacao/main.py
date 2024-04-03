
import openpyxl
import os
from random import uniform

#criando um unico arquivo que tem duas planilhas

caminho_base = os.path.dirname(os.path.realpath(__file__))
caminho_novo_arquivo = caminho_base + r'\nova_planilha.xlsx'


planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1',0)
planilha.create_sheet('Planilha2',1)


planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']


for linha in range(1,11):  
    numero_pedido = linha - 1
    planilha1.cell(linha,1).value = numero_pedido   
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10 ,100),2)
    planilha1.cell(linha, 3).value = preco


for linha in range(1,11):  
    planilha2.cell(linha,1).value = f'nome {linha} {round(uniform(10,100),2)}'
    planilha2.cell(linha,2).value = f'sobrenome {linha} {round(uniform(10,100),2)}'
    planilha2.cell(linha,3).value = f'idade {linha} {round(uniform(10,100),2)}'


planilha.save(caminho_novo_arquivo)

