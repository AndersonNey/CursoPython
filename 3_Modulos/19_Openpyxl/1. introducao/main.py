"""
https://openpyxl.readthedocs.io/en/stable/
pip install openpyxl
pipenv install openpyxl
"""

import openpyxl
import os
from random import uniform

caminho_base = os.path.dirname(os.path.realpath(__file__))
caminho_arquivo = caminho_base + r'\pedidos.xlsx'
caminho_novo_arquivo = caminho_base + r'\nova_planilha.xlsx'


pedidos = openpyxl.load_workbook(caminho_arquivo)
nome_planilhas = pedidos.sheetnames    #retorna todas as planilhas dentro do documento
planilha1 = pedidos['Página1']

#---OS CODIGOD ABAIXO FORAM TESTADOS AQUI------------

for linha in range(5,16):  #salvando os dados apartir da linha 5 ate a 15
    numero_pedido = linha - 1
    planilha1.cell(linha,1).value = numero_pedido   #esse .cell() é aqui que passo as coordenadas da celula em numeros tanto a linha quanto a coluna
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10 ,100),2)
    planilha1.cell(linha, 3).value = preco

#----------------------------------------------------
pedidos.save(caminho_novo_arquivo) #salvando os dados na nova_planilha




#dicas 

# 1 acessando uma coluna-----------------------------------------------------------------------------------------

# for campo in planilha1['b']:
#     print(campo.value)

# 2 acessando um campo--------------------------------------------------------------------------------------------

# print(planilha1['b4'].value)

# 3 acessando ranges-----------------------------------------------------------------------------------------------


# for linha in planilha1['a1:c2']:
#     for coluna in linha:
#         print(coluna.value)

# 4 acessando a planilha inteira----------------------------------------------------------------------------------

# for linha in planilha1:
#     for coluna in linha:
#         print(coluna.value)

# 5 Organizar os dados da planilha para exibir bonito--------------------------------------------------------------

# for linha in planilha1:
#     print(linha[0].value,linha[1].value,linha[2].value,linha[3].value)

# 6 trocando o valor de um campo-----------------------------------------------------------------------------------

# planilha1['b3'].value = 2000

# 7 acrescntando mais dados a tabela-------------------------------------------------------------------------------

# for linha in range(5,16):  #salvando os dados apartir da linha 5 ate a 15
#     numero_pedido = linha - 1
#     planilha1.cell(linha,1).value = numero_pedido   #esse .cell() é aqui que passo as coordenadas da celula em numeros tanto a linha quanto a coluna
#     planilha1.cell(linha, 2).value = 1200 + linha

#     preco = round(uniform(10 ,100),2)
#     planilha1.cell(linha, 3).value = preco





